"""
Excel → PostgreSQL Converter  🇲🇽
Motor de lectura : DuckDB + openpyxl
UI               : CustomTkinter
Tema             : Bandera de México (Verde · Blanco · Rojo)

Instalar:
    pip install customtkinter duckdb openpyxl tkinterdnd2

─────────────────────────────────────────────────────────
CORRECCIONES DE RENDIMIENTO (v3):
  1. Lectura en HILO SEPARADO → la UI nunca se congela
  2. Auto-detección de columna real máxima (evita ZZ=702 cols)
  3. Centinela de filas vacías: para al encontrar 5 filas
     vacías consecutivas (evita iterar 1M filas fantasma)
  4. Barra de progreso + botón Cancelar
  5. Rango inteligente: si header_row>1 y rango vacío,
     sólo salta filas, NO genera rango "A2:ZZ9999999"
─────────────────────────────────────────────────────────
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import re
import os
import tempfile
import csv as _csv
import threading
from pathlib import Path
from datetime import datetime, date

try:
    import duckdb
    DUCKDB_OK = True
except ImportError:
    DUCKDB_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Paleta México ──────────────────────────────────────────────────────────────
VERDE = "#006847"
VERDE_L = "#00855C"
BLANCO = "#FFFFFF"
ROJO = "#CE1126"
ROJO_L = "#E8192C"
GRIS = "#F0F0F0"
GRIS_L = "#E8E8E8"
TEXTO = "#1A1A1A"
AMBAR = "#CC8800"

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES DE RANGO
# ═══════════════════════════════════════════════════════════════════════════════

def col_letter_to_index(col: str) -> int:
    r = 0
    for c in col.upper():
        r = r * 26 + (ord(c) - ord('A') + 1)
    return r - 1


def index_to_col_letter(idx: int) -> str:
    """0-based index → Excel column letter(s)."""
    result = ""
    idx += 1
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def parse_range(range_str: str):
    """Devuelve (r0,c0,r1,c1) 0-based o None si vacío."""
    s = range_str.strip().upper()
    if not s:
        return None
    m = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", s)
    if not m:
        raise ValueError(f"Rango inválido: '{range_str}'. Usa formato A1:D10")
    c1, r1, c2, r2 = m.groups()
    r0 = int(r1) - 1
    c0 = col_letter_to_index(c1)
    r1e = int(r2) - 1 if r2 else r0
    c1e = col_letter_to_index(c2) if c2 else c0
    return r0, c0, r1e, c1e


def _sanitize_headers(raw_headers: list) -> list:
    seen: dict = {}
    result = []
    for i, h in enumerate(raw_headers):
        name = str(h).strip() if h is not None else f"col_{i}"
        name = re.sub(r'[^\w]+', '_', name,
                      flags=re.UNICODE).strip('_') or f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        result.append(name)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  LECTURA OPTIMIZADA CON OPENPYXL  (el cuello de botella real)
# ═══════════════════════════════════════════════════════════════════════════════

def get_sheet_names(filepath: str) -> list:
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no instalado: pip install openpyxl")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _read_sheet_optimized(
    filepath: str,
    sheet_name: str,
    rng,            # None o (r0,c0,r1,c1) 0-based — si None: auto
    header_row: bool,
    cancel_flag,
    progress_cb,
    max_empty: int = 5,
) -> tuple:
    """
    Lee la hoja con tres optimizaciones clave:
      1. max_col = ws.max_column REAL (nunca 702 columnas fantasma)
      2. Centinela de filas vacías (para en max_empty filas seguidas)
      3. Cancel via threading.Event
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    reported_max = ws.max_row or 0
    # FIX CRÍTICO: usar max_column REAL de la hoja, cap 500 por seguridad
    real_max_col = min(ws.max_column or 30, 500)

    if rng is not None:
        r0, c0, r1, c1 = rng
        min_row = r0 + 1
        max_row = min(r1 + 1, reported_max) if r1 < 5_000_000 else reported_max
        min_col = c0 + 1
        max_col = min(c1 + 1, real_max_col) if c1 < 5_000 else real_max_col
    else:
        min_row = 1
        max_row = reported_max
        min_col = 1
        max_col = real_max_col

    progress_cb(
        2, f"'{sheet_name}': {reported_max:,} filas × {real_max_col} cols")

    data = []
    empty_streak = 0
    total_hint = min(max(reported_max - min_row + 1, 1), 200_000)

    for i, row in enumerate(ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True
    )):
        if cancel_flag.is_set():
            wb.close()
            return None, None

        row_list = list(row)
        has_data = any(v is not None for v in row_list)

        if has_data:
            empty_streak = 0
            data.append(row_list)
        else:
            empty_streak += 1
            if empty_streak >= max_empty:
                break
            if data:
                data.append(row_list)

        if i % 500 == 0:
            pct = min(90, int(i / total_hint * 85)) if total_hint > 0 else 10
            progress_cb(pct, f"Leyendo fila {min_row + i:,}…")

    wb.close()
    while data and all(v is None for v in data[-1]):
        data.pop()

    if not data:
        return [], []

    if header_row:
        return data[0], data[1:]
    else:
        return [None] * len(data[0]), data


# ═══════════════════════════════════════════════════════════════════════════════
#  MOTOR DUCKDB
# ═══════════════════════════════════════════════════════════════════════════════

def read_excel_with_duckdb(
    filepath: str, sheet_name: str, range_str: str,
    header_row: bool, header_row_num: int,
    cancel_flag, progress_cb
):
    """
    Pipeline: openpyxl (optimizado) → CSV temporal → DuckDB → DuckDBDataFrame
    """
    if not DUCKDB_OK:
        raise RuntimeError("duckdb no instalado: pip install duckdb")
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no instalado: pip install openpyxl")

    rng = parse_range(range_str) if range_str.strip() else None

    # Si header_row > 1 y no hay rango explícito: saltar filas de encabezado
    # Usamos max_col=5000 como valor grande → _read_sheet_optimized lo recorta
    # al ws.max_column real, evitando iterar columnas vacías.
    if rng is None and header_row_num >= 2:
        rng = (header_row_num - 1, 0, 5_000_000, 5_000)

    progress_cb(1, "Iniciando lectura…")
    raw_headers, rows = _read_sheet_optimized(
        filepath, sheet_name, rng, header_row, cancel_flag, progress_cb
    )

    if raw_headers is None:   # cancelado
        return None

    if not rows and not raw_headers:
        return DuckDBDataFrame([], {})

    headers = _sanitize_headers(raw_headers)
    n_cols = len(headers)

    progress_cb(88, "Escribiendo CSV temporal…")

    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False,
        encoding="utf-8-sig", newline=""
    )
    try:
        writer = _csv.writer(tmp)
        writer.writerow(headers)
        for row in rows:
            if cancel_flag.is_set():
                tmp.close()
                os.unlink(tmp.name)
                return None
            padded = (list(row) + [None] * n_cols)[:n_cols]
            writer.writerow(["" if v is None else str(v) for v in padded])
        tmp.close()

        progress_cb(93, "DuckDB infiriendo tipos…")
        safe_path = tmp.name.replace("\\", "/")
        con = duckdb.connect()
        rel = con.execute(
            f"SELECT * FROM read_csv_auto('{safe_path}', "
            f"header=true, sample_size=-1, null_padding=true)"
        ).fetchdf()
        col_data = {c: rel[c].tolist() for c in rel.columns}
        raw_dtypes = {c: str(rel[c].dtype) for c in rel.columns}
        con.close()

        progress_cb(100, "✅ Listo")
        return DuckDBDataFrame(list(rel.columns), col_data, raw_dtypes)

    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


# ── DuckDBDataFrame ────────────────────────────────────────────────────────────

class DuckDBDataFrame:
    _PG = {
        "int8": "SMALLINT", "int16": "SMALLINT", "int32": "INTEGER",
        "int64": "BIGINT",  "uint8": "SMALLINT", "uint16": "INTEGER",
        "uint32": "BIGINT", "uint64": "NUMERIC",
        "float16": "REAL",  "float32": "REAL",   "float64": "DOUBLE PRECISION",
        "bool": "BOOLEAN",  "boolean": "BOOLEAN",
        "object": "TEXT",   "string": "TEXT",
        "datetime64[ns]": "TIMESTAMP", "datetime64[us]": "TIMESTAMP",
        "timedelta64[ns]": "INTERVAL", "date32[day]": "DATE",
    }

    def __init__(self, columns, data, raw_dtypes=None):
        self.columns = list(columns)
        self._data = data
        self._raw = raw_dtypes or {}

    @property
    def shape(self):
        if not self.columns:
            return (0, 0)
        return (len(self._data[self.columns[0]]), len(self.columns))

    def is_empty(self):
        return self.shape[0] == 0 or not self.columns

    def iter_rows(self):
        n = self.shape[0]
        for i in range(n):
            yield tuple(self._data[c][i] for c in self.columns)

    def pg_type(self, col: str) -> str:
        raw = self._raw.get(col, "object").lower()
        if raw in self._PG:
            return self._PG[raw]
        for k, v in self._PG.items():
            if raw.startswith(k):
                return v
        return "TEXT"

    def describe_schema(self) -> str:
        lines = ["SCHEMA  —  Motor: DuckDB\n", "=" * 55 + "\n\n"]
        lines.append(f"  {'Columna':<36} {'DuckDB':<20} PostgreSQL\n")
        lines.append("  " + "─" * 70 + "\n")
        for c in self.columns:
            lines.append(
                f"  {c:<36} {self._raw.get(c, '?'):<20} {self.pg_type(c)}\n")
        n, m = self.shape
        lines += [f"\n{'=' * 55}\n", f"Filas: {n:,}  |  Columnas: {m}\n"]
        return "".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
#  GENERADOR DDL POSTGRESQL
# ═══════════════════════════════════════════════════════════════════════════════

def _safe_col(c: str) -> str:
    return re.sub(r'\W+', '_', c).strip('_').lower() or "col"


def duckdb_to_postgresql_ddl(df: DuckDBDataFrame, table_name: str) -> str:
    safe_table = re.sub(
        r'\W+', '_', table_name).strip('_').lower() or "mi_tabla"
    col_defs = [f'    "{_safe_col(c)}" {df.pg_type(c)}' for c in df.columns]
    col_names = ", ".join(f'"{_safe_col(c)}"' for c in df.columns)

    ddl = "-- ✅ Generado por Excel→PostgreSQL Converter  (motor: DuckDB)\n"
    ddl += f"-- Tabla: {safe_table}  |  {df.shape[0]:,} filas × {df.shape[1]} cols\n\n"
    ddl += f'DROP TABLE IF EXISTS "{safe_table}";\n'
    ddl += f'CREATE TABLE "{safe_table}" (\n'
    ddl += ",\n".join(col_defs) + "\n);\n\n"

    if df.is_empty():
        return ddl

    BATCH = 500
    buf = []
    batch_n = 0

    def flush(buf, is_last):
        nonlocal ddl, batch_n
        batch_n += 1
        ddl += f'INSERT INTO "{safe_table}" ({col_names}) VALUES\n'
        ddl += ",\n".join(buf) + ";\n"
        if not is_last:
            ddl += f"\n-- lote {batch_n}\n"

    # Cadenas que representan valores nulos en pandas/DuckDB
    _NULL_STRINGS = frozenset({
        "nat", "nan", "none", "null", "<na>", "n/a", "na", "", "pd.nat",
    })

    for i, row in enumerate(df.iter_rows()):
        vals = []
        for v in row:
            # ── Nulos nativos ──────────────────────────────────────────────
            if v is None:
                vals.append("NULL")
                continue
            # ── NaN float (pandas produce estos) ──────────────────────────
            if isinstance(v, float) and v != v:
                vals.append("NULL")
                continue
            # ── NaT y otros nulos como string (pandas → DuckDB → lista) ──
            str_v = str(v).strip()
            if str_v.lower() in _NULL_STRINGS:
                vals.append("NULL")
                continue
            # ── Booleano ──────────────────────────────────────────────────
            if isinstance(v, bool):
                vals.append("TRUE" if v else "FALSE")
            # ── Numérico ──────────────────────────────────────────────────
            elif isinstance(v, (int, float)):
                vals.append(str(v))
            # ── Fecha / Timestamp ─────────────────────────────────────────
            elif isinstance(v, (datetime, date)):
                vals.append(f"'{v}'")
            # ── Texto ─────────────────────────────────────────────────────
            else:
                vals.append("'" + str_v.replace("'", "''") + "'")
        buf.append("  (" + ", ".join(vals) + ")")
        if len(buf) == BATCH:
            flush(buf, i + 1 == df.shape[0])
            buf = []

    if buf:
        flush(buf, True)
    return ddl


# ═══════════════════════════════════════════════════════════════════════════════
#  INTERFAZ GRÁFICA
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelToPGApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("🇲🇽  Excel → PostgreSQL  |  Motor: DuckDB")
        self.geometry("1160x820")
        self.minsize(960, 680)
        self.configure(fg_color=BLANCO)

        self.filepath = None
        self.df_actual = None
        self._cancel = threading.Event()
        self._reading = False

        self._build_ui()

        try:
            from tkinterdnd2 import DND_FILES
            self.drop_zone.drop_target_register(DND_FILES)
            self.drop_zone.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass

    # ── Construcción UI ───────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        ctk.CTkFrame(self, fg_color=ROJO, height=5,
                     corner_radius=0).pack(fill="x")
        body = ctk.CTkFrame(self, fg_color=BLANCO, corner_radius=0)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1, minsize=310)
        body.columnconfigure(1, weight=3)
        body.rowconfigure(0, weight=1)
        self._build_left(body)
        self._build_right(body)
        ctk.CTkFrame(self, fg_color=VERDE, height=3,
                     corner_radius=0).pack(fill="x", side="bottom")
        self._build_statusbar()

    def _build_header(self):
        h = ctk.CTkFrame(self, fg_color=VERDE, height=60, corner_radius=0)
        h.pack(fill="x")
        h.pack_propagate(False)
        ctk.CTkLabel(h,
                     text="🇲🇽  Excel → PostgreSQL Converter",
                     font=ctk.CTkFont(family="Arial", size=21, weight="bold"),
                     text_color=BLANCO
                     ).pack(side="left", padx=22)
        badge = ctk.CTkFrame(h, fg_color=ROJO, corner_radius=6)
        badge.pack(side="right", padx=16, pady=14)
        ctk.CTkLabel(badge,
                     text="  ⚡ DuckDB  ",
                     font=ctk.CTkFont(family="Arial", size=11, weight="bold"),
                     text_color=BLANCO
                     ).pack(padx=4, pady=2)

    def _build_statusbar(self):
        sb = ctk.CTkFrame(self, fg_color=GRIS_L, height=30, corner_radius=0)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Listo.")
        self.status_lbl = ctk.CTkLabel(
            sb, textvariable=self.status_var,
            font=ctk.CTkFont(family="Arial", size=11),
            text_color=VERDE, anchor="w"
        )
        self.status_lbl.pack(side="left", padx=12, pady=4)

    # ── Panel izquierdo ───────────────────────────────────────────────────────
    def _build_left(self, parent):
        panel = ctk.CTkFrame(parent, fg_color=GRIS, corner_radius=0)
        panel.grid(row=0, column=0, sticky="nsew")
        inner = ctk.CTkScrollableFrame(panel, fg_color=GRIS, corner_radius=0)
        inner.pack(fill="both", expand=True, padx=10, pady=10)

        # Drop zone
        self.drop_zone = ctk.CTkFrame(inner, fg_color=BLANCO,
                                      border_width=2, border_color=VERDE, corner_radius=14, height=115)
        self.drop_zone.pack(fill="x", pady=(0, 6))
        self.drop_zone.pack_propagate(False)
        self.drop_icon = ctk.CTkLabel(self.drop_zone, text="📂",
                                      font=ctk.CTkFont(family="Arial", size=28))
        self.drop_icon.pack(pady=(12, 2))
        self.drop_label = ctk.CTkLabel(self.drop_zone,
                                       text="Arrastra Excel aquí  o  haz clic",
                                       font=ctk.CTkFont(
                                           family="Arial", size=12),
                                       text_color=VERDE, justify="center")
        self.drop_label.pack()
        for w in (self.drop_zone, self.drop_icon, self.drop_label):
            w.bind("<Button-1>", lambda e: self._browse())

        self.file_label = ctk.CTkLabel(inner, text="",
                                       font=ctk.CTkFont(
                                           family="Arial", size=10),
                                       text_color="#666", wraplength=260, justify="left")
        self.file_label.pack(fill="x", pady=(2, 8))

        # Hoja
        self._sec(inner, "📋  Hoja de Excel")
        self.sheet_var = tk.StringVar(value="")
        self.sheet_menu = ctk.CTkOptionMenu(inner, variable=self.sheet_var,
                                            values=[
                                                "(carga un archivo primero)"],
                                            fg_color=VERDE, button_color=VERDE_L, button_hover_color=ROJO,
                                            text_color=BLANCO, dropdown_fg_color=BLANCO,
                                            dropdown_text_color=TEXTO, dropdown_hover_color=GRIS_L,
                                            command=self._on_sheet_change)
        self.sheet_menu.pack(fill="x", pady=(2, 8))

        # Fila encabezado
        self._sec(inner, "📌  Fila de encabezados")
        hf = ctk.CTkFrame(inner, fg_color=GRIS)
        hf.pack(fill="x", pady=(2, 8))
        ctk.CTkLabel(hf, text="Fila #",
                     font=ctk.CTkFont(family="Arial", size=11),
                     text_color=TEXTO).pack(side="left", padx=(0, 6))
        self.header_row_var = tk.IntVar(value=1)
        tk.Spinbox(hf, from_=0, to=50, width=5,
                   textvariable=self.header_row_var,
                   font=("Arial", 11), relief="flat",
                   bg=BLANCO, fg=TEXTO,
                   highlightthickness=1, highlightcolor=VERDE,
                   buttonbackground=GRIS
                   ).pack(side="left")
        ctk.CTkLabel(hf, text="  (0 = sin encabezado)",
                     font=ctk.CTkFont(family="Arial", size=10),
                     text_color="#888").pack(side="left")

        # Rango
        self._sec(inner, "📐  Rango de celdas")
        ctk.CTkLabel(inner,
                     text="Ej: A2:Z5000  (vacío = auto)",
                     font=ctk.CTkFont(family="Arial", size=10),
                     text_color="#888").pack(anchor="w")
        self.range_entry = ctk.CTkEntry(inner,
                                        placeholder_text="vacío = detección automática",
                                        border_color=VERDE, fg_color=BLANCO)
        self.range_entry.pack(fill="x", pady=(2, 8))

        # Nombre tabla
        self._sec(inner, "🗄️  Tabla PostgreSQL")
        self.table_entry = ctk.CTkEntry(inner,
                                        placeholder_text="nombre_tabla",
                                        border_color=VERDE, fg_color=BLANCO)
        self.table_entry.pack(fill="x", pady=(2, 14))

        # ── Barra de progreso ─────────────────────────────────────────────────
        self.progress_bar = ctk.CTkProgressBar(inner,
                                               fg_color=GRIS_L,
                                               progress_color=VERDE,
                                               height=10, corner_radius=5)
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", pady=(0, 4))

        self.progress_lbl = ctk.CTkLabel(inner, text="",
                                         font=ctk.CTkFont(
                                             family="Arial", size=10),
                                         text_color="#666")
        self.progress_lbl.pack(anchor="w", pady=(0, 8))

        # Botones
        self.btn_read = ctk.CTkButton(inner,
                                      text="🔍  Leer con DuckDB",
                                      fg_color=VERDE, hover_color=VERDE_L, text_color=BLANCO,
                                      font=ctk.CTkFont(
                                          family="Arial", size=13, weight="bold"),
                                      height=42, corner_radius=10,
                                      command=self._read_excel)
        self.btn_read.pack(fill="x", pady=(0, 6))

        self.btn_cancel = ctk.CTkButton(inner,
                                        text="⛔  Cancelar lectura",
                                        fg_color="#888", hover_color=ROJO, text_color=BLANCO,
                                        font=ctk.CTkFont(
                                            family="Arial", size=11),
                                        height=32, corner_radius=8,
                                        state="disabled",
                                        command=self._cancel_read)
        self.btn_cancel.pack(fill="x", pady=(0, 6))

        ctk.CTkButton(inner,
                      text="⚡  Generar SQL PostgreSQL",
                      fg_color=ROJO, hover_color=ROJO_L, text_color=BLANCO,
                      font=ctk.CTkFont(family="Arial", size=13, weight="bold"),
                      height=42, corner_radius=10,
                      command=self._generate_sql
                      ).pack(fill="x", pady=(0, 6))

        ctk.CTkButton(inner,
                      text="💾  Exportar .sql",
                      fg_color="#444", hover_color="#222", text_color=BLANCO,
                      font=ctk.CTkFont(family="Arial", size=11),
                      height=34, corner_radius=8,
                      command=self._export_sql
                      ).pack(fill="x")

        # Estado dependencias
        ok = DUCKDB_OK and OPENPYXL_OK
        missing = ", ".join(
            ([] if DUCKDB_OK else ["duckdb"]) +
            ([] if OPENPYXL_OK else ["openpyxl"])
        )
        ctk.CTkLabel(inner,
                     text="✅ DuckDB + openpyxl listos" if ok else f"⚠️ Falta: {missing}",
                     font=ctk.CTkFont(family="Arial", size=10),
                     text_color=VERDE if ok else ROJO, justify="left"
                     ).pack(anchor="w", pady=(14, 0))

    # ── Panel derecho ─────────────────────────────────────────────────────────
    def _build_right(self, parent):
        panel = ctk.CTkFrame(parent, fg_color=BLANCO, corner_radius=0)
        panel.grid(row=0, column=1, sticky="nsew")

        tabs = ctk.CTkTabview(panel,
                              fg_color=BLANCO,
                              segmented_button_fg_color=GRIS,
                              segmented_button_selected_color=VERDE,
                              segmented_button_selected_hover_color=VERDE_L,
                              segmented_button_unselected_color=GRIS,
                              segmented_button_unselected_hover_color=GRIS_L,
                              text_color=TEXTO)
        tabs.pack(fill="both", expand=True, padx=10, pady=10)
        tabs.add("📊  Vista Previa")
        tabs.add("🐘  SQL PostgreSQL")
        tabs.add("ℹ️  Schema / Tipos")

        # Vista previa
        prev = tabs.tab("📊  Vista Previa")
        ft = ctk.CTkFrame(prev, fg_color=BLANCO, corner_radius=8,
                          border_width=1, border_color=GRIS_L)
        ft.pack(fill="both", expand=True, padx=4, pady=4)
        sty = ttk.Style()
        sty.theme_use("default")
        sty.configure("MX.Treeview",
                      background=BLANCO, foreground=TEXTO,
                      fieldbackground=BLANCO, rowheight=24, font=("Arial", 10))
        sty.configure("MX.Treeview.Heading",
                      background=VERDE, foreground=BLANCO,
                      font=("Arial", 10, "bold"), relief="flat")
        sty.map("MX.Treeview",
                background=[("selected", VERDE_L)],
                foreground=[("selected", BLANCO)])
        self.tree = ttk.Treeview(ft, style="MX.Treeview", show="headings")
        vsb = ttk.Scrollbar(ft, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(ft, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.info_label = ctk.CTkLabel(prev, text="",
                                       font=ctk.CTkFont(
                                           family="Arial", size=11),
                                       text_color="#555", anchor="w")
        self.info_label.pack(anchor="w", padx=8, pady=4)

        # SQL
        sql_tab = tabs.tab("🐘  SQL PostgreSQL")
        self.sql_text = ctk.CTkTextbox(sql_tab,
                                       fg_color="#1A1A2E", text_color="#E0E0E0",
                                       font=ctk.CTkFont(
                                           family="Courier New", size=11),
                                       corner_radius=8, wrap="none")
        self.sql_text.pack(fill="both", expand=True, padx=4, pady=(4, 2))
        br = ctk.CTkFrame(sql_tab, fg_color=BLANCO)
        br.pack(fill="x", padx=4, pady=(0, 4))
        ctk.CTkButton(br, text="📋 Copiar",
                      fg_color=VERDE, hover_color=VERDE_L, text_color=BLANCO,
                      width=110, height=30, command=self._copy_sql
                      ).pack(side="left", padx=4)
        ctk.CTkButton(br, text="🗑️ Limpiar",
                      fg_color=ROJO, hover_color=ROJO_L, text_color=BLANCO,
                      width=90, height=30,
                      command=lambda: self.sql_text.delete("1.0", "end")
                      ).pack(side="left", padx=4)

        # Schema
        sch = tabs.tab("ℹ️  Schema / Tipos")
        self.schema_text = ctk.CTkTextbox(sch,
                                          fg_color="#F8F8F8", text_color=TEXTO,
                                          font=ctk.CTkFont(
                                              family="Courier New", size=11),
                                          corner_radius=8)
        self.schema_text.pack(fill="both", expand=True, padx=4, pady=4)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _sec(self, p, title):
        ctk.CTkLabel(p, text=title,
                     font=ctk.CTkFont(family="Arial", size=12, weight="bold"),
                     text_color=VERDE
                     ).pack(anchor="w", pady=(8, 2))

    def _status(self, msg: str, color: str = VERDE):
        self.status_var.set(msg)
        self.status_lbl.configure(text_color=color)

    def _set_progress(self, pct: int, msg: str = ""):
        """Llamado desde el hilo de lectura via after()."""
        self.progress_bar.set(pct / 100)
        if msg:
            self.progress_lbl.configure(text=msg)
            self._status(msg, AMBAR if pct < 100 else VERDE)

    def _progress_cb(self, pct: int, msg: str):
        """Thread-safe: schedula actualización en el hilo principal."""
        self.after(0, lambda: self._set_progress(pct, msg))

    # ── Eventos ───────────────────────────────────────────────────────────────
    def _browse(self):
        p = filedialog.askopenfilename(
            title="Selecciona archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls *.xlsm"), ("Todos", "*.*")]
        )
        if p:
            self._load_file(p)

    def _on_drop(self, event):
        p = event.data.strip().strip("{}")
        if p.lower().endswith((".xlsx", ".xls", ".xlsm")):
            self._load_file(p)
        else:
            messagebox.showwarning("Formato inválido",
                                   "Solo se aceptan archivos Excel (.xlsx .xls .xlsm)")

    def _on_sheet_change(self, _=None):
        if not self.table_entry.get().strip():
            s = re.sub(r'\W+', '_', self.sheet_var.get()).strip('_').lower()
            self.table_entry.delete(0, "end")
            self.table_entry.insert(0, s)

    def _load_file(self, path: str):
        try:
            sheets = get_sheet_names(path)
            self.filepath = path
            short = Path(path).name
            self.file_label.configure(text=f"📄 {short}")
            self.drop_label.configure(text=f"✅ {short}", text_color=ROJO)
            self.drop_icon.configure(text="📗")
            self.drop_zone.configure(border_color=ROJO)
            self.sheet_menu.configure(values=sheets)
            self.sheet_var.set(sheets[0])
            self._on_sheet_change()
            self._status(f"✅ {short}  —  {len(sheets)} hoja(s)", VERDE)
            self.progress_bar.set(0)
            self.progress_lbl.configure(text="")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self._status(f"❌ {e}", ROJO)

    def _cancel_read(self):
        self._cancel.set()
        self._status("⛔ Cancelando…", ROJO)
        self.btn_cancel.configure(state="disabled")

    def _read_excel(self):
        if self._reading:
            return
        if not self.filepath:
            messagebox.showwarning(
                "Sin archivo", "Primero carga un archivo Excel.")
            return
        if not DUCKDB_OK:
            messagebox.showerror("DuckDB no instalado",
                                 "Ejecuta:\n  pip install duckdb")
            return

        sheet = self.sheet_var.get()
        range_str = self.range_entry.get().strip()
        h_row_num = self.header_row_var.get()
        use_header = h_row_num >= 1

        # Bloquear UI de lectura
        self._reading = True
        self._cancel.clear()
        self.btn_read.configure(state="disabled", text="⏳ Leyendo…")
        self.btn_cancel.configure(state="normal")
        self.progress_bar.set(0)
        self.progress_lbl.configure(text="Iniciando…")

        def worker():
            try:
                df = read_excel_with_duckdb(
                    self.filepath, sheet, range_str,
                    use_header, h_row_num,
                    self._cancel, self._progress_cb
                )
                self.after(0, lambda: self._on_read_done(df, sheet))
            except Exception as e:
                self.after(0, lambda err=e: self._on_read_error(err))

        threading.Thread(target=worker, daemon=True).start()

    def _on_read_done(self, df, sheet: str):
        self._reading = False
        self.btn_read.configure(state="normal", text="🔍  Leer con DuckDB")
        self.btn_cancel.configure(state="disabled")

        if df is None:
            self._status("⛔ Lectura cancelada", ROJO)
            self.progress_bar.set(0)
            self.progress_lbl.configure(text="Cancelado")
            return

        self.df_actual = df
        self._populate_tree(df)
        self._populate_schema(df)
        n, m = df.shape
        self._status(
            f"✅ DuckDB leyó {n:,} filas × {m} columnas  —  hoja: '{sheet}'",
            VERDE
        )

    def _on_read_error(self, err: Exception):
        self._reading = False
        self.btn_read.configure(state="normal", text="🔍  Leer con DuckDB")
        self.btn_cancel.configure(state="disabled")
        messagebox.showerror("Error de lectura", str(err))
        self._status(f"❌ {err}", ROJO)
        self.progress_bar.set(0)
        self.progress_lbl.configure(text="Error")

    def _generate_sql(self):
        if self.df_actual is None:
            messagebox.showwarning("Sin datos", "Primero lee el Excel.")
            return
        table = self.table_entry.get().strip() or "mi_tabla"
        try:
            sql = duckdb_to_postgresql_ddl(self.df_actual, table)
            self.sql_text.delete("1.0", "end")
            self.sql_text.insert("1.0", sql)
            self._status(f"✅ SQL generado  —  tabla: '{table}'", VERDE)
        except Exception as e:
            messagebox.showerror("Error SQL", str(e))
            self._status(f"❌ {e}", ROJO)

    def _export_sql(self):
        sql = self.sql_text.get("1.0", "end").strip()
        if not sql:
            messagebox.showwarning("Sin SQL", "Primero genera el SQL.")
            return
        table = self.table_entry.get().strip() or "exportacion"
        # Mostrar diálogo de configuración PostgreSQL antes de guardar
        dlg = PgExportDialog(self, table_name=table, sql=sql)
        self.wait_window(dlg)

    def _copy_sql(self):
        self.clipboard_clear()
        self.clipboard_append(self.sql_text.get("1.0", "end"))
        self._status("📋 SQL copiado al portapapeles", VERDE)

    def _populate_tree(self, df: DuckDBDataFrame):
        self.tree.delete(*self.tree.get_children())
        if df.is_empty():
            self.info_label.configure(text="Sin datos")
            return
        cols = df.columns
        self.tree.configure(columns=cols)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=max(
                90, min(200, len(c) * 9)), minwidth=60)
        MAX = 500
        for i, row in enumerate(df.iter_rows()):
            tag = "odd" if i % 2 else "even"
            self.tree.insert("", "end",
                             values=[("" if v is None else str(v))
                                     for v in row],
                             tags=(tag,))
            if i >= MAX - 1:
                break
        self.tree.tag_configure("odd",  background="#F0FFF4")
        self.tree.tag_configure("even", background=BLANCO)
        n, m = df.shape
        self.info_label.configure(
            text=f"Total: {n:,} filas × {m} columnas" +
            (f"  (preview: primeras {MAX})" if n > MAX else ""))

    def _populate_schema(self, df: DuckDBDataFrame):
        self.schema_text.delete("1.0", "end")
        self.schema_text.insert("1.0", df.describe_schema())


# ═══════════════════════════════════════════════════════════════════════════════
#  DIÁLOGO DE EXPORTACIÓN / CONEXIÓN POSTGRESQL
# ═══════════════════════════════════════════════════════════════════════════════

class PgExportDialog(ctk.CTkToplevel):
    """
    Ventana emergente para:
      1. Configurar datos de conexión PostgreSQL
      2. Elegir nombre de archivo .sql de destino
      3. Opcionalmente añadir cabecera de conexión al SQL
    Defaults tomados de la instancia pgAdmin del usuario.
    """

    def __init__(self, parent, table_name: str, sql: str):
        super().__init__(parent)
        self.parent = parent
        self.sql = sql
        self.title("💾  Exportar SQL — Configuración PostgreSQL")
        self.geometry("560x700")
        self.minsize(520, 680)
        self.resizable(True, True)
        self.configure(fg_color=BLANCO)
        self.grab_set()
        self.focus_set()
        self.lift()
        self.after(50, self._center)
        self._build(table_name)

    def _center(self):
        self.update_idletasks()
        pw = self.parent.winfo_rootx()
        ph = self.parent.winfo_rooty()
        pw2 = self.parent.winfo_width()
        ph2 = self.parent.winfo_height()
        x = pw + (pw2 - 560) // 2
        y = max(30, ph + (ph2 - 700) // 2)
        self.geometry(f"560x700+{x}+{y}")

    def _build(self, table_name: str):
        # ── Header ─────────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=VERDE, height=52, corner_radius=0)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr,
                     text="🐘  Exportar a PostgreSQL",
                     font=ctk.CTkFont(family="Arial", size=16, weight="bold"),
                     text_color=BLANCO
                     ).pack(side="left", padx=18, pady=10)

        ctk.CTkFrame(self, fg_color=ROJO, height=4,
                     corner_radius=0).pack(fill="x", side="top")

        # ── Botones: SIEMPRE al fondo (pack side=bottom antes que el body) ────
        footer = ctk.CTkFrame(self, fg_color=BLANCO, corner_radius=0)
        footer.pack(fill="x", side="bottom", padx=0, pady=0)

        ctk.CTkFrame(footer, fg_color=GRIS_L, height=1,
                     corner_radius=0).pack(fill="x")
        btn_row = ctk.CTkFrame(footer, fg_color=BLANCO, corner_radius=0)
        btn_row.pack(fill="x", padx=18, pady=12)

        ctk.CTkButton(btn_row,
                      text="✖  Cancelar",
                      fg_color="#888", hover_color="#555", text_color=BLANCO,
                      width=110, height=40, corner_radius=8,
                      command=self.destroy
                      ).pack(side="right", padx=(6, 0))

        ctk.CTkButton(btn_row,
                      text="💾  Guardar .sql",
                      fg_color=VERDE, hover_color=VERDE_L, text_color=BLANCO,
                      width=155, height=40, corner_radius=8,
                      font=ctk.CTkFont(family="Arial", size=13, weight="bold"),
                      command=self._save
                      ).pack(side="right", padx=(6, 0))

        ctk.CTkButton(btn_row,
                      text="📋  Copiar psql",
                      fg_color=ROJO, hover_color=ROJO_L, text_color=BLANCO,
                      width=130, height=40, corner_radius=8,
                      font=ctk.CTkFont(family="Arial", size=11),
                      command=self._copy_conn
                      ).pack(side="left")

        # ── Cuerpo scrollable (DESPUÉS de los botones para que queden fijos) ──
        body = ctk.CTkScrollableFrame(self, fg_color=BLANCO, corner_radius=0)
        body.pack(fill="both", expand=True, padx=18, pady=(8, 0))

        def field(parent, label, default, show=""):
            ctk.CTkLabel(parent, text=label,
                         font=ctk.CTkFont(
                             family="Arial", size=11, weight="bold"),
                         text_color=VERDE, anchor="w"
                         ).pack(fill="x", pady=(8, 1))
            e = ctk.CTkEntry(parent, fg_color=GRIS, border_color=VERDE,
                             border_width=1, show=show)
            e.insert(0, default)
            e.pack(fill="x")
            return e

        ctk.CTkLabel(body,
                     text="Datos de conexión PostgreSQL",
                     font=ctk.CTkFont(family="Arial", size=13, weight="bold"),
                     text_color=TEXTO
                     ).pack(anchor="w", pady=(0, 4))

        ctk.CTkFrame(body, fg_color=GRIS_L, height=1).pack(
            fill="x", pady=(0, 6))

        # Campos con defaults de tu pgAdmin
        row1 = ctk.CTkFrame(body, fg_color=BLANCO)
        row1.pack(fill="x")
        row1.columnconfigure(0, weight=3)
        row1.columnconfigure(1, weight=1)

        left1 = ctk.CTkFrame(row1, fg_color=BLANCO)
        left1.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        right1 = ctk.CTkFrame(row1, fg_color=BLANCO)
        right1.grid(row=0, column=1, sticky="ew")

        self.e_host = field(left1,  "🌐  Host",  "localhost")
        self.e_port = field(right1, "🔌  Puerto", "5432")

        self.e_db = field(body, "🗄️  Base de datos", "postgres")
        self.e_user = field(body, "👤  Usuario",       "postgres")
        self.e_pass = field(body, "🔑  Contraseña",    "", show="●")
        self.e_tbl = field(body, "📋  Nombre de tabla", table_name)

        ctk.CTkFrame(body, fg_color=GRIS_L, height=1).pack(
            fill="x", pady=(14, 6))

        # Opciones
        ctk.CTkLabel(body,
                     text="Opciones de exportación",
                     font=ctk.CTkFont(family="Arial", size=13, weight="bold"),
                     text_color=TEXTO
                     ).pack(anchor="w", pady=(0, 6))

        self.chk_header = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(body,
                        text="Añadir cabecera con datos de conexión al archivo SQL",
                        variable=self.chk_header,
                        fg_color=VERDE, hover_color=VERDE_L, checkmark_color=BLANCO,
                        font=ctk.CTkFont(family="Arial", size=11), text_color=TEXTO
                        ).pack(anchor="w", pady=(0, 4))

        self.chk_drop = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(body,
                        text="Incluir DROP TABLE IF EXISTS antes del CREATE",
                        variable=self.chk_drop,
                        fg_color=VERDE, hover_color=VERDE_L, checkmark_color=BLANCO,
                        font=ctk.CTkFont(family="Arial", size=11), text_color=TEXTO
                        ).pack(anchor="w", pady=(0, 16))

        # ── Botón principal visible siempre dentro del scroll ─────────────────
        ctk.CTkFrame(body, fg_color=GRIS_L, height=1).pack(
            fill="x", pady=(0, 12))

        ctk.CTkButton(body,
                      text="💾  Guardar archivo .sql",
                      fg_color=VERDE, hover_color=VERDE_L, text_color=BLANCO,
                      height=48, corner_radius=10,
                      font=ctk.CTkFont(family="Arial", size=14, weight="bold"),
                      command=self._save
                      ).pack(fill="x", pady=(0, 6))

        ctk.CTkButton(body,
                      text="🗃️  Guardar como .db (SQLite)",
                      fg_color="#2C5F8A", hover_color="#1E4060", text_color=BLANCO,
                      height=36, corner_radius=8,
                      font=ctk.CTkFont(family="Arial", size=11),
                      command=self._save_sqlite
                      ).pack(fill="x", pady=(0, 8))

        ctk.CTkButton(body,
                      text="📋  Copiar comando psql al portapapeles",
                      fg_color=ROJO, hover_color=ROJO_L, text_color=BLANCO,
                      height=36, corner_radius=8,
                      font=ctk.CTkFont(family="Arial", size=11),
                      command=self._copy_conn
                      ).pack(fill="x", pady=(0, 4))

        ctk.CTkButton(body,
                      text="✖  Cancelar",
                      fg_color="#888", hover_color="#555", text_color=BLANCO,
                      height=32, corner_radius=8,
                      font=ctk.CTkFont(family="Arial", size=11),
                      command=self.destroy
                      ).pack(fill="x", pady=(0, 12))

    def _build_header_comment(self) -> str:
        # Genera cabecera compatible con pgAdmin Query Tool (SQL puro).
        # NO incluye meta-comandos psql (como backslash-connect) que rompen pgAdmin.
        # El comando psql completo se muestra como comentario informativo.
        db = self.e_db.get()
        host = self.e_host.get()
        port = self.e_port.get()
        user = self.e_user.get()
        tbl = self.e_tbl.get()
        psql_cmd = f"psql -h {host} -p {port} -U {user} -d {db} -f archivo.sql"
        return (
            f"-- ═══════════════════════════════════════════════════════════\n"
            f"-- Generado por Excel→PostgreSQL Converter (motor: DuckDB)\n"
            f"-- ───────────────────────────────────────────────────────────\n"
            f"--   Host      : {host}\n"
            f"--   Puerto    : {port}\n"
            f"--   Base datos: {db}\n"
            f"--   Usuario   : {user}\n"
            f"--   Tabla     : {tbl}\n"
            f"-- ───────────────────────────────────────────────────────────\n"
            f"-- Para ejecutar desde terminal:\n"
            f"--   {psql_cmd}\n"
            f"-- ═══════════════════════════════════════════════════════════\n\n"
        )

    def _build_psql_cmd(self) -> str:
        host = self.e_host.get()
        port = self.e_port.get()
        db = self.e_db.get()
        user = self.e_user.get()
        return f"psql -h {host} -p {port} -U {user} -d {db} -f <archivo.sql>"

    def _get_final_sql(self) -> str:
        sql = self.sql
        # Ajustar nombre de tabla si cambió
        new_table = re.sub(r'\W+', '_', self.e_tbl.get()
                           ).strip('_').lower() or "mi_tabla"
        # Quitar DROP TABLE si el usuario no lo quiere
        if not self.chk_drop.get():
            sql = re.sub(r'DROP TABLE IF EXISTS "[^"]+";\n', '', sql)
        # Añadir cabecera de conexión
        final = ""
        if self.chk_header.get():
            final += self._build_header_comment()
        final += sql
        return final

    def _save(self):
        table = self.e_tbl.get().strip() or "exportacion"
        p = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".sql",
            filetypes=[("SQL", "*.sql"), ("Texto", "*.txt")],
            initialfile=f"{table}.sql",
            title="Guardar archivo SQL"
        )
        if not p:
            return
        try:
            Path(p).write_text(self._get_final_sql(), encoding="utf-8")
            # Mostrar resumen en un popup de confirmación
            info = (
                f"✅ Archivo guardado:\n{Path(p).name}\n\n"
                f"Para ejecutar en PostgreSQL:\n"
                f"{self._build_psql_cmd()}"
            )
            messagebox.showinfo("Exportación exitosa", info, parent=self)
            self.parent._status(f"💾 Exportado: {Path(p).name}", VERDE)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e), parent=self)

    def _save_sqlite(self):
        """
        Exporta el DataFrame actual a un archivo .db SQLite.
        Usa sqlite3 estándar (sin dependencias extra).
        Funciona independientemente del SQL generado para PostgreSQL.
        """
        if self.parent.df_actual is None or self.parent.df_actual.is_empty():
            messagebox.showwarning(
                "Sin datos", "No hay datos cargados.", parent=self)
            return

        table = re.sub(r'\W+', '_', self.e_tbl.get()
                       ).strip('_').lower() or "datos"
        p = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".db",
            filetypes=[("SQLite DB", "*.db"),
                       ("SQLite DB", "*.sqlite"), ("Todos", "*.*")],
            initialfile=f"{table}.db",
            title="Guardar base de datos SQLite"
        )
        if not p:
            return
        try:
            import sqlite3
            df = self.parent.df_actual

            def _safe(c):
                return re.sub(r'\W+', '_', c).strip('_').lower() or "col"

            # Mapear tipos DuckDB → SQLite
            _SQLITE = {
                "int": "INTEGER", "float": "REAL", "double": "REAL",
                "bool": "INTEGER", "text": "TEXT", "object": "TEXT",
                "date": "TEXT", "time": "TEXT",
            }

            def sqlite_type(col):
                raw = df._raw.get(col, "object").lower()
                for k, v in _SQLITE.items():
                    if raw.startswith(k):
                        return v
                return "TEXT"

            col_defs = ", ".join(
                f'"{_safe(c)}" {sqlite_type(c)}' for c in df.columns)
            col_names = ", ".join(f'"{_safe(c)}"' for c in df.columns)

            _NULL_STR = frozenset(
                {"nat", "nan", "none", "null", "<na>", "n/a", "na", "", "pd.nat"})

            con = sqlite3.connect(p)
            cur = con.cursor()
            cur.execute(f'DROP TABLE IF EXISTS "{table}"')
            cur.execute(f'CREATE TABLE "{table}" ({col_defs})')

            for row in df.iter_rows():
                vals = []
                for v in row:
                    if v is None or str(v).strip().lower() in _NULL_STR:
                        vals.append(None)
                    elif isinstance(v, float) and v != v:
                        vals.append(None)
                    else:
                        vals.append(str(v))
                cur.execute(
                    f'INSERT INTO "{table}" ({col_names}) VALUES ({",".join(["?"]*len(vals))})',
                    vals
                )
            con.commit()
            con.close()

            n, m = df.shape
            messagebox.showinfo(
                "SQLite exportado",
                f"✅ Base de datos guardada:\n{Path(p).name}\n\n"
                f"{n:,} filas × {m} columnas → tabla '{table}'\n\n"
                f"Ábrela con:\n  DB Browser for SQLite\n  DBeaver\n  TablePlus",
                parent=self
            )
            self.parent._status(f"🗃️ SQLite guardado: {Path(p).name}", VERDE)
        except Exception as e:
            messagebox.showerror("Error SQLite", str(e), parent=self)

    def _copy_conn(self):
        cmd = self._build_psql_cmd()
        self.parent.clipboard_clear()
        self.parent.clipboard_append(cmd)
        self.parent._status("📋 Comando psql copiado al portapapeles", VERDE)


# ═══════════════════════════════════════════════════════════════════════════════
def main():
    app = ExcelToPGApp()
    app.mainloop()


if __name__ == "__main__":
    main()
